"""
Build python models using parameter values from Excel.

| variable | module | distribution      | param 1 | param 2 | param 3 | unit | label                      |
|----------|--------|-------------------|---------|---------|---------|------|----------------------------|
| c_ON     | numpy.random  | triangular | 0.25    | 0.75    | 1       | -    | number WDM terminals metro |

"""
import logging
from collections import defaultdict

import importlib
import itertools
import operator
from openpyxl.utils.datetime import from_excel
import xlrd

from excel_helper import growth_coefficients
from openpyxl.utils.datetime import from_excel
import pandas as pd
import numpy as np

from dateutil import relativedelta as rdelta
from xarray import Dataset, DataArray

SINGLE_VAR = 'single_var'

logger = logging.getLogger(__name__)

__author__ = 'schien'

NAME = 'name'
DISTRIBUTION = 'distribution'
SCENARIO = 'scenario'
PARAM_A = 'param_a'
PARAM_B = 'param_b'
PARAM_C = 'param_c'
MODULE = 'module'
LABEL = 'label'
COMMENT = 'comment'
START_DATE = 'start date'
END_DATE = 'end date'
CAGR = 'CAGR'
REF_DATE = 'ref date'
SOURCE = 'source'
UNIT = 'unit'

HEADER_SEQ = [NAME, SCENARIO, MODULE, DISTRIBUTION, PARAM_A, PARAM_B, PARAM_C, UNIT, START_DATE, END_DATE, CAGR,
              REF_DATE, LABEL, COMMENT, SOURCE]

TABLE_STRUCT = {k: i for i, k in enumerate(HEADER_SEQ)}
INV_TABLE_STRUCT = {v: k for k, v in TABLE_STRUCT.items()}
DEFAULT_SCENARIO = 'def'

from functools import total_ordering


@total_ordering
class MinType(object):
    """
    Needed to make None orderable in python 3.
    """

    def __le__(self, other):
        return True

    def __eq__(self, other):
        return (self is other)


Min = MinType()


def get_analytical_mean(options, X):
    name = options['name']
    if name == 'normal':
        return options['params'][0]
    if name == 'uniform':
        return (options['params'][0] + options['params'][1]) / 2
    if name == 'choice':
        return options['params'][0]
    if name == 'triangular':
        return (options['params'][0] + options['params'][1] + options['params'][2]) / 3
    return X.mean()


class ParameterLoader(object):
    """
    Load parameters from excel or a tabular source.
    Init variables from distribution values.

    A cache is used for the distribution configuration. However, every time get_val or __getitem__ is called,
    a new vector of random numbers is created.
    """

    @classmethod
    def from_excel(cls, file, size=1, sheet_index=None, sheet_name=None, **kwargs):
        """Initialize from an excel file"""
        rows = load_workbook(file, sheet_index, sheet_name)

        return cls(rows, size, **kwargs)

    def __init__(self, rows, size, **kwargs):
        """
        Rows is a list of tuples.

        :param rows:
        :param size:
        :param kwargs:
        """
        self.kwargs = kwargs
        self.data = defaultdict(list)

        sorted_wb = sorted(rows, key=lambda x: Min if x[1] is None else x[1])
        for key, group in itertools.groupby(sorted_wb, operator.itemgetter(1)):

            # scenario = row[1] if row[1] else DEFAULT_SCENARIO
            if not key:
                key = DEFAULT_SCENARIO
            self.data[key] = list(group)

            self.size = size
            self.cache = defaultdict(dict)

        self.scenario = DEFAULT_SCENARIO

        self.sample_mean_value = kwargs.get('sample_mean_value', False)

    def select_scenario(self, scenario):
        self.scenario = scenario

    def unselect_scenario(self):
        self.scenario = DEFAULT_SCENARIO

    def get_row(self, name):
        scenario = self.scenario
        variables = [row[TABLE_STRUCT[NAME]] for row in self.data[scenario]]
        i = variables.index(name)
        return self.data[scenario][i]

    def generate_values(self, f, p, options, size=None, name=None):
        """
        Instantiates the random variable.

        :param size:
        :param f:
        :param p:
        :param options:
        :param kwargs:
        :return:
        """
        if not size:  # can be overridden case by case
            size = self.size

        # generate the distribution
        X = f(*p, size=size)
        if type(X) is not float:
            X = X.astype('f')

        if self.sample_mean_value:
            return np.full(size, get_analytical_mean(options, X))

        if SINGLE_VAR in self.kwargs:

            # use it for all variables, other than the single var
            # if this is the single var, make sure 'with_single_var_cagr' is set to True
            if not self.kwargs[SINGLE_VAR] == name or \
                            hasattr(self, 'with_single_var_cagr') and \
                            not self.with_single_var_cagr:
                # reduce to mean
                return np.full(size, get_analytical_mean(options, X))

        return X

    def get_val(self, name):
        """
        Generate a random variable as is defined in the source
        If a variable definition for scenario exists, use that.


        args: optional additonal args
        If no args are given, applies default size from constructor
        """
        scenario = self.scenario
        f, p = None, None
        if name in self.cache[scenario]:
            f, p, options = self.cache[scenario][name]
        else:
            row = self.get_row(name)
            f, p, options = get_random_variable_definition(row)
            self.cache[scenario][name] = (f, p, options)

        ret = self.generate_values(f, p, options, name=name)

        return ret

    def __contains__(self, item):

        return item in self.cache or item in [tup[0] for tup in self.data[self.scenario]]

    def clear_cache(self):
        self.cache = defaultdict(dict)

    def get_label(self, name):
        try:
            row = self.get_row(name)
        except:
            return name
        return row[TABLE_STRUCT[LABEL]]

    def get_property(self, name, prop):
        try:
            row = self.get_row(name)
        except:
            return name
        return row[TABLE_STRUCT[prop]]

    def __getitem__(self, name):
        """
        Get the distribution for a item name from the table
        Then execute and return the result array
        """
        return self.get_val(name)


def load_workbook(file, sheet_index=None, sheet_name=None):
    """
    Build a list of tuples from the columns of the excel table.

    :param file:
    :return: a list of tuples
    [
        (col1,col2,...),
        (col1,col2,...),
    ]
    """
    wb = xlrd.open_workbook(file)
    sh = None
    if sheet_name is not None:
        sh = wb.sheet_by_name(sheet_name)
    if sheet_index is not None:
        sh = wb.sheet_by_index(sheet_index)
    if not sh:
        raise Exception('Must provide either sheet name or sheet index')

    def transform(col_idx, col):

        # strip white spaces
        if INV_TABLE_STRUCT[col_idx] in [NAME]:
            return [i.strip() for i in col]
        # transform dates
        if INV_TABLE_STRUCT[col_idx] in [START_DATE, END_DATE, REF_DATE]:
            return [from_excel(i) if type(i) == float else i for i in col]
        return col

    rows_es = zip(*[transform(col_idx, sh.col_values(col_idx)[1:]) for col_idx in range(0, len(HEADER_SEQ))])
    return rows_es


def get_random_variable_definition(row):
    """

    :param row:
    :return:
    """

    module = importlib.import_module(row[TABLE_STRUCT[MODULE]])
    distribution_name = row[TABLE_STRUCT[DISTRIBUTION]]
    func = getattr(module, distribution_name)
    if distribution_name == 'choice':

        cell = row[TABLE_STRUCT[PARAM_A]]
        if type(cell) in [float, int]:
            params = ([cell],)
        else:
            tokens = cell.split(',')

            params = [float(token.strip()) for token in tokens]
            params = (params,)
    # @todo rename to 'empirical'
    elif distribution_name == 'Distribution':
        func = func()
        params = tuple(row[TABLE_STRUCT[i]] for i in [PARAM_A, PARAM_B, PARAM_C] if row[TABLE_STRUCT[i]])
    else:
        params = tuple(
            row[TABLE_STRUCT[i]] for i in [PARAM_A, PARAM_B, PARAM_C] if row[TABLE_STRUCT[i]] not in [None, ''])

    options = {key: row[TABLE_STRUCT[key]] for key in [UNIT, LABEL, COMMENT, START_DATE, END_DATE, CAGR, REF_DATE] if
               row[TABLE_STRUCT[key]]}

    options['name'] = distribution_name
    options['params'] = params

    return func, params, options


def growth_coefficients(start_date, end_date, ref_date, alpha, samples):
    """
    Build a matrix of growth factors according to the CAGR formula  y'=y0 (1+a)^(t'-t0).
    The
    """
    if ref_date < start_date:
        raise ValueError("Ref data must be >= start date.")
    if ref_date > end_date:
        raise ValueError("Ref data must be >= start date.")
    if ref_date > start_date and alpha >= 1:
        raise ValueError("For a CAGR >= 1, ref date and start date must be the same.")
    # relative delta will be positive if ref date is >= start date (which it should with above assertions)
    delta_ar = rdelta.relativedelta(ref_date, start_date)
    ar = delta_ar.months + 12 * delta_ar.years
    delta_br = rdelta.relativedelta(end_date, ref_date)
    br = delta_br.months + 12 * delta_br.years

    # we place the ref point on the lower interval (delta_ar + 1) but let it start from 0
    # in turn we let the upper interval start from 1
    g = np.fromfunction(lambda i, j: np.power(1 - alpha, np.abs(i) / 12), (ar + 1, samples), dtype=float)
    h = np.fromfunction(lambda i, j: np.power(1 + alpha, np.abs(i + 1) / 12), (br, samples), dtype=float)
    g = np.flipud(g)
    # now join the two arrays
    return np.vstack((g, h))


class DataSeriesLoader(ParameterLoader):
    """
    A dataloader that supports timelines.
    """

    @classmethod
    def from_excel(cls, file, times, size=1, sheet_index=None, sheet_name=None, index_names=['time', 'samples'],
                   **kwargs):
        """Initialize from an excel file
        :param file:
        :param times:
        :param size:
        :param sheet_index:
        :param sheet_name:
        :param index_names:
        :param kwargs:
        :return:
        """
        rows = load_workbook(file, sheet_index, sheet_name)
        return cls(rows, times, index_names, size, **kwargs)

    @classmethod
    def from_dataframe(cls, df, times, size=1, index_names=['time', 'samples'], **kwargs):
        return cls(df.where((pd.notnull(df)), None).values, times, index_names, size, **kwargs)

    def __init__(self, rows, times, index_names, size, **kwargs):
        super(DataSeriesLoader, self).__init__(rows, size, **kwargs)
        self.kwargs = kwargs
        # single var variability is on by default
        self.with_single_var_cagr = kwargs['with_single_var_cagr'] if 'with_single_var_cagr' in kwargs else True

        self.times = times
        iterables = [times, range(0, size)]
        self._multi_index = pd.MultiIndex.from_product(iterables, names=index_names)
        assert type(times.freq) == pd.tseries.offsets.MonthBegin, 'Time index must have monthly frequency'

    def get_val(self, name):
        """
        Wraps the value in a dataframe
        :param name:
        :param kwargs:
        :return:
        """
        df = super(DataSeriesLoader, self).get_val(name)
        df.set_index(self._multi_index, inplace=True)
        df.columns = [name]
        # @todo this is a hack to return a series with index as I don't know how to set an index and rename a series
        data_series = df.ix[:, 0]
        data_series._metadata = df._metadata
        return data_series

    def generate_values(self, f, p, options, size=None, name=None):
        """
        Instantiate a random variable and apply annual growth factors.


        :param name:
        :param size:
        :param f:
        :param p:
        :param options:

        :return:
        """
        logger.info('generating values for variable %s' % name)
        values = super(DataSeriesLoader, self).generate_values(f, p, options, size=(len(self.times) * self.size,),
                                                               name=name)
        alpha = 0
        # apply CAGR if present - or if we are in 'single_var mode' then, this needs to be the single_var
        if CAGR in options.keys() and REF_DATE in options.keys() and \
                (SINGLE_VAR not in self.kwargs or self.with_single_var_cagr and self.kwargs[SINGLE_VAR] == name):
            alpha = options[CAGR]

        # @todo - fill to cover the entire time: define rules for filling first
        ref_date = options[REF_DATE] if REF_DATE in options else self.times[0].to_pydatetime()
        assert ref_date >= self.times[0].to_pydatetime(), 'Ref date must be within variable time span.'
        assert ref_date <= self.times[-1].to_pydatetime(), 'Ref date must be within variable time span.'

        start_date = self.times[0].to_pydatetime()
        end_date = self.times[-1].to_pydatetime()

        a = growth_coefficients(start_date, end_date, ref_date, alpha, self.size)

        values *= a.ravel()

        df = pd.DataFrame(values)
        df._metadata = [options]
        return df


from abc import ABCMeta, abstractmethod


class LoaderDataSource(object):
    """
    A handle to files or other stores with variable definitions
    such that we can reload them when the underlying file has changed.
    """
    __metaclass__ = ABCMeta

    def __init__(self, size):
        self.size = size

    @abstractmethod
    def get_loader(self, **kwargs):
        pass


class ExcelLoaderMixin(object):
    def __key(self):
        return (self.file, self.sheet_name, self.sheet_index)

    def __eq__(self, y):
        return self.__key() == y.__key()

    def __hash__(self):
        return hash(self.__key())

    def __repr__(self):
        return "%s (file: %s, sheet name: %s, sheet index: %s)" % (
            self.__class__, self.file, self.sheet_name, self.sheet_index)


class ExcelLoaderDataSource(ExcelLoaderMixin, LoaderDataSource):
    """
    Source for one-dimensional random variables
    """

    def __init__(self, file, size=1, sheet_index=None, sheet_name=None):
        super(ExcelLoaderDataSource, self).__init__(size)

        self.sheet_name = sheet_name
        self.sheet_index = sheet_index
        self.file = file

    def get_loader(self, **kwargs):
        return ParameterLoader.from_excel(self.file, self.size, self.sheet_index, self.sheet_name, **kwargs)


class DataSeriesLoaderDataSource(LoaderDataSource):
    """
    Abstract class that has a times property for time series data.
    This data is a two-dimensional random variable with multiple values for each time step.
    """
    __metaclass__ = ABCMeta

    def __init__(self, times, size=1):
        super(DataSeriesLoaderDataSource, self).__init__(size)
        self.times = times

    @abstractmethod
    def get_loader(self, **kwargs):
        return super(DataSeriesLoaderDataSource, self).get_loader(**kwargs)


# @todo - this is a terrible hack, iron this out...
loader_cache = defaultdict(dict)


class ExcelSeriesLoaderDataSource(ExcelLoaderMixin, DataSeriesLoaderDataSource):
    """
    Load data from excel.
    """

    def __init__(self, file, times, size=1, sheet_index=None, sheet_name=None, index_names=None, **kwargs):
        super(ExcelSeriesLoaderDataSource, self).__init__(times, size)
        if index_names is None:
            index_names = ['time', 'samples']
        self.index_names = index_names
        self.sheet_name = sheet_name
        self.sheet_index = sheet_index
        self.file = file
        self.kwargs = kwargs

    def get_loader(self, **kwargs):

        if not self.file in loader_cache or self.sheet_name not in loader_cache[self.file]:
            loader_cache[self.file][self.sheet_name] = DataSeriesLoader.from_excel(self.file, self.times, self.size,
                                                                                   self.sheet_index, self.sheet_name,
                                                                                   self.index_names, **kwargs,
                                                                                   **self.kwargs)

        return loader_cache[self.file][self.sheet_name]


class DataFrameLoaderDataSource(DataSeriesLoaderDataSource):
    """
    Load variables from a dataframe.

    Needed to comply with the interface.
    Does not actually do much.
    """

    def __init__(self, name, df, times, size=1, index_names=None):
        super(DataFrameLoaderDataSource, self).__init__(times, size)
        if index_names is None:
            index_names = ['time', 'samples']
        self.name = name
        self.index_names = index_names
        self.df = df

    def __key(self):
        return (self.name)

    def __eq__(self, y):
        return self.__key() == y.__key()

    def __hash__(self):
        return hash(self.__key())

    def get_loader(self, **kwargs):
        return DataSeriesLoader.from_dataframe(self.df, self.times, self.size, self.index_names, **kwargs)


class MultiSourceLoader(object):
    """
    Bundles sources loaders.
    A variable can be instantiated from any of the underlying sources in a first come, first serve manner.
    """

    def __init__(self, **kwargs):
        self._sources = {}
        self.kwargs = kwargs

    def add_source(self, source):
        """
        Add a source to load variable definitions from.
        Sources are of type ::
        :param source:
        :return:
        """
        loader = source.get_loader(**self.kwargs)
        self._sources[source] = loader

    def __getitem__(self, item):
        for loader in self._sources.values():
            if item in loader:
                return loader[item]

    def set_scenario(self, scenario):
        for loader in self._sources.values():
            loader.select_scenario(scenario)

    def reset_scenario(self):
        for loader in self._sources.values():
            loader.unselect_scenario()

    def reload_sources(self):
        for source, loader in self._sources.items():
            logger.info('Reloading source %s' % source)
            self._sources[source] = source.get_loader(**self.kwargs)


class MCDataset(Dataset):
    """
    Wraps an xray.Dataset such that variables that are not found in the dataset are first looked up in the
    SourceLoader dict.
    """

    def __init__(self, **kwargs):
        self.known_items = []
        self._ldr = MultiSourceLoader(**kwargs)
        super(MCDataset, self).__init__()

    def add_source(self, loader):
        logger.info('Adding source loader %s' % loader)
        assert isinstance(loader, LoaderDataSource)
        self._ldr.add_source(loader)

    def reload_sources(self):
        logger.info('Reloading sources')
        self._ldr.reload_sources()

    def prepare(self, item):
        """
        Load a variable from the data source into the :xray.Dataset:

        :param item:
        :return:
        """

        series = self._ldr[item]
        if series is None:
            raise LookupError('%s not found in any of the loaders' % item)
        meta = series._metadata[0]
        # take the abs because we don't want negative values
        s = DataArray.from_series(series.abs())
        s.attrs.update(meta)
        self[item] = s
        self.known_items.append(item)
        return s

    def __getitem__(self, item):
        """
        Return a variable from the backing :xray.Dataset:
        If the variable is not present in the Dataset it is instantiated from a source.
        :param item:
        :return:
        """
        if item == '_ipython_canary_method_should_not_exist_':
            print('trying to load %s' % item)
            return
        try:

            # print('trying to load %s' % item)
            return super(MCDataset, self).__getitem__(item)
        except Exception as inst:
            # print('%s not found, trying to load' % item)
            self.prepare(item)
            return super(MCDataset, self).__getitem__(item)

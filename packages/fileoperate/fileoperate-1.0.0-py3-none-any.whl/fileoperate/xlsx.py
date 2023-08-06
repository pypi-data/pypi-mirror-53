import openpyxl
from robotLibrary.common.log import ILog

logger = ILog(__file__)


def creatExcel(file, sheetName) :
    '''
    @功能 创建一个excel文件 \n
    @参数 file 文件名， sheetName sheet页名 \n
    @返回 如果成功返回True，失败抛出异常 \n
    '''
    # logger.info('creatExcel方法执行开始')
    # logger.info('传入参数:file='+str(file)+'sheetName='+str(sheetName))
    try :
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = str(sheetName)
        wb.save(file)
        return True
    except Exception as e :
        logger.exception('creatExcel方法执行出现异常')
    finally :
        logger.info('creatExcel方法执行结束wb=' + str(wb))


def loadWorkBook(file) :
    '''
    @功能 加载excel文件  \n
    @参数 file excel全路径 \n
    @返回 Exception \n
    '''
    # logger.info('loadWorkBook方法执行开始')
    # logger.info('传入参数:file='+str(file))
    try :
        wb = openpyxl.load_workbook(file)
        # logger.info('excel加载成功')
    except Exception as e :
        logger.exception('loadWorkBook方法执行出现异常')
    finally :
        logger.info('loadWorkBook方法执行结束wb=' + str(wb))
    return wb


def getSheet(wb, sheetName) :
    '''
    @功能 获取sheet页  \n
    @参数 wb 工作簿 \n
    @参数 sheetName sheet页名称 \n
    @返回 Exception \n
    '''
    # logger.info('getSheet方法执行开始')
    # logger.info('传入参数:wb='+str(wb)+'sheetName='+str(sheetName))
    try :
        sheet = wb[sheetName]
        # logger.info('sheet页对象获取成功')
    except Exception as e :
        logger.exception('getSheet方法执行出现异常')
    finally :
        logger.info('getSheet方法执行结束sheet=' + str(sheet))
    return sheet


def getMaxRow(sheet) :
    '''
    @功能 获取当前sheet页最大行  \n
    @参数 sheet sheet页 \n
    @返回 Exception \n
    '''
    # logger.info('getMaxRow方法执行开始')
    # logger.info('传入参数:sheet='+str(sheet))
    try :
        max_row = sheet.max_row
        # logger.info('sheet页最大行获取成功')
    except Exception as e :
        logger.exception('getMaxRow方法执行出现异常')
    finally :
        logger.info('getMaxRow方法执行结束sheet=' + str(sheet))
    return max_row


def getCellData(sheet, row, column) :
    '''
    @功能 获取单元格数据  \n
    @参数 sheet sheet页 \n
    @参数 row 行号 \n
    @参数 column 列号 \n
    @返回 Exception \n
    '''
    # logger.info('getCellData方法执行开始')
    # logger.info('传入参数:sheet='+str(sheet)+'row='+str(row)+'column='+str(column))
    try :
        data = str(sheet.cell(row=int(row), column=int(column)).value)
        # logger.info('单元格数据获取成功')
    except Exception as e :
        logger.exception('getCellData方法执行出现异常')
    finally :
        logger.debug('getCellData方法执行结束data=' + str(data))
    return data


def setCellData(sheet, row, column, value) :
    '''
    @功能 设置单元格值  \n
    @参数 sheet sheet页 \n
    @参数 row 行号 \n
    @参数 column 列号 \n
    @参数 value 值 \n
    @返回 Exception \n
    '''
    # logger.info('setCellData方法执行开始')
    # logger.info('传入参数:sheet='+str(sheet)+'row='+str(row)+'column='+str(column))
    try :
        sheet.cell(row=int(row), column=int(column), value=str(value))
        # logger.info('单元格数据获取成功')
    except Exception as e :
        logger.exception('setCellData方法执行出现异常')
    finally :
        logger.info('setCellData方法执行结束data=' + str(value))
    return value


def mergeCells(sheet, start_row, start_column, end_row, end_column) :
    '''
    @功能 拆分单元格  \n
    @参数 sheet sheet页 \n
    @参数 start_row 开始行号 \n
    @参数 start_column 开始列号 \n
    @参数 end_row 结束行号 \n
    @参数 end_column 结束列号 \n
    @返回 Exception \n
    '''
    # logger.info('mergeCells方法执行开始')
    # logger.info('传入参数:sheet='+str(sheet)+'start_row='+str(start_row)+'start_column='+str(start_column)+'end_row='+str(end_row)+'end_column='+str(end_column))
    try :
        sheet.merge_cells(start_row=int(start_row), start_column=int(start_column), end_row=int(end_row),
                          end_column=int(end_column))
        # logger.info('合并单元格成功')
    except Exception as e :
        logger.exception('mergeCells方法执行出现异常')
    finally :
        logger.info('mergeCells方法执行结束')


def cellAlignment(sheet, row, column, horizontal=None, vertical=None, border_style=None, border_color=None,
                  font_size=None, font_color=None, fill_color=None, fill_type=None, wrapText=None) :
    '''
    @功能 设置单元格样式  \n
    @参数 sheet sheet页 \n
    @参数 row 行号 \n
    @参数 column 列号 \n
    @参数 horizontal 水平方向 horizontal_alignments取值说明:('general','left','center','right','fill','justify','centerContinuous','distributed') \n
    @参数 vertical 垂直方向 vertical_aligments取值说明:('top', 'center', 'bottom', 'justify','distributed') \n
    @参数 border_style 单元格样式 border_style取值说明:('dashDot','dashDotDot', 'dashed','dotted','double','hair', 'medium', 'mediumDashDot', 'mediumDashDotDot','mediumDashed', 'slantDashDot', 'thick', 'thin') \n
    @参数 border_color 单元格颜色 \n
    @参数 font_size 字体大小 \n
    @参数 font_color 字体颜色 \n
    @参数 fill_color 填充颜色 \n
    @参数 fill_type 填充类型 fill_type取值说明:'gray0625', 'lightHorizontal', 'lightVertical', 'gray125', 'darkVertical', 'darkGray', 'darkDown', 'darkTrellis', 'lightTrellis', 'lightDown', 'darkGrid', 'lightUp', 'lightGrid', 'mediumGray', 'solid', 'darkUp', 'darkHorizontal', 'lightGray' \n
    @返回 Exception \n
    例子：cellAlignment(sheet1,i,j,horizontal='center',vertical='center',border_style='medium',border_color='E53528',font_size="22",font_color='0864B1',fill_color='F4CF15',fill_type='solid')
    '''
    # 例子：cellAlignment(sheet1,i,j,horizontal='center',vertical='center',border_style='medium',border_color='E53528',font_size="22",font_color='0864B1',fill_color='F4CF15',fill_type='solid')
    # logger.info('cellAlignment方法执行开始')
    # logger.info('传入参数:sheet='+str(sheet)+'row='+str(row)+'column='+str(column)+'horizontal='+str(horizontal)+'vertical='+str(vertical)+'border_style='+str(border_style)+'border_color='+str(border_color)+'font_size='+str(font_size)+'font_color='+str(font_color)+'fill_color='+str(fill_color)+'fill_type='+str(fill_type))
    try :
        if (font_size != None and str(font_size).strip() != '') or (
                font_color != None and str(font_color).strip() != '') :
            font = openpyxl.styles.Font(size=font_size, color=font_color)
            sheet.cell(row=row, column=column).font = font
        if (horizontal != None and str(horizontal).strip() != '') or (
                vertical != None and str(vertical).strip() != '') :
            ##logger.info("horizontal_alignments取值说明:('general','left','center','right','fill','justify','centerContinuous','distributed')")
            ##logger.info("vertical_aligments取值说明:('top', 'center', 'bottom', 'justify','distributed')")
            alignment = openpyxl.styles.Alignment(horizontal=horizontal, vertical=vertical, wrapText=wrapText)
            sheet.cell(row=row, column=column).alignment = alignment
        if (fill_color != None and str(fill_color).strip() != '') or (
                fill_type != None and str(fill_type).strip() != '') :
            ##logger.info("fill_type取值说明:'gray0625', 'lightHorizontal', 'lightVertical', 'gray125', 'darkVertical', 'darkGray', 'darkDown', 'darkTrellis', 'lightTrellis', 'lightDown', 'darkGrid', 'lightUp', 'lightGrid', 'mediumGray', 'solid', 'darkUp', 'darkHorizontal', 'lightGray'")
            fill = openpyxl.styles.PatternFill(start_color=fill_color, end_color=fill_color, fill_type=fill_type)
            sheet.cell(row=row, column=column).fill = fill
        if (border_style != None and str(border_style).strip() != '') or (
                border_color != None and str(border_color).strip() != '') :
            ##logger.info("border_style取值说明:('dashDot','dashDotDot', 'dashed','dotted','double','hair', 'medium', 'mediumDashDot', 'mediumDashDotDot','mediumDashed', 'slantDashDot', 'thick', 'thin')")
            border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style=border_style, color=border_color),
                right=openpyxl.styles.Side(style=border_style, color=border_color),
                top=openpyxl.styles.Side(style=border_style, color=border_color),
                bottom=openpyxl.styles.Side(style=border_style, color=border_color),
                diagonal=openpyxl.styles.Side(style=border_style, color=border_color),
                diagonal_direction=0,
                outline=openpyxl.styles.Side(style=border_style, color=border_color),
                vertical=openpyxl.styles.Side(style=border_style, color=border_color),
                horizontal=openpyxl.styles.Side(style=border_style, color=border_color)
            )
            sheet.cell(row=row, column=column).border = border
        # logger.info('单元格样式调整成功')
    except Exception as e :
        logger.exception('cellAlignment方法执行出现异常')
    finally :
        logger.info('mergeCells方法执行结束')


def saveExcel(wb, file) :
    '''
    @功能 保存excel  \n
    @参数 wb 工作簿 \n
    @参数 file excel全路径 \n
    @返回 Exception \n
    '''
    # logger.info('saveExcel方法执行开始')
    # logger.info('传入参数:wb='+str(wb)+'file='+str(file))
    try :
        wb.save(file)
        # #logger.info('excel保存成功')
    except Exception as e :
        logger.exception('saveExcel方法执行出现异常')
    finally :
        logger.info('saveExcel方法执行结束wb=' + str(wb))


def closeExcel(wb, file) :
    '''
    @功能 关闭excel  \n
    @参数 wb 工作簿 \n
    @参数 file excel全路径 \n
    @返回 Exception \n
    '''
    # logger.info('closeExcel方法执行开始')
    # logger.info('传入参数:wb='+str(wb)+'file='+str(file))
    try :
        saveExcel(wb, file)
        wb.close()
        # logger.info('Excel关闭成功')
    except Exception as e :
        logger.exception('closeExcel方法执行出现异常')
    finally :
        logger.info('closeExcel方法执行结束wb=' + str(wb))

# 操作excel样例
# wb = loadWorkBook(r"C:\Users\LLS\Desktop\text.xlsx")
# sheet1 = getSheet(wb,"Sheet1")
# for i in range(1,15):
#    for j in range(1,6):
#        cellAlignment(sheet1,i,j,horizontal='center',vertical='center',border_style='medium',border_color='E53528',font_size="22",font_color='0864B1',fill_color='F4CF15',fill_type='solid')
# closeExcel(wb,r"C:\Users\LLS\Desktop\text.xlsx")

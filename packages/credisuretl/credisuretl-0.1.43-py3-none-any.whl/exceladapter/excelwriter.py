import openpyxl


class ExcelWriter:
    def __init__(self, filename):
        self.workbook = openpyxl.Workbook()
        self.filename = filename

    def create_sheet(self, sheetname):
        return self.workbook.create_sheet(index=0, title=sheetname)

    def build_sheet(self, sheetname, sheet_data):
        sheet = self.workbook.create_sheet(index=0, title=sheetname)
        for key, value in sheet_data.items():
            sheet[key] = value

    def save(self):
        self.workbook.save(filename=self.filename)

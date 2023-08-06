import xlrd
from openpyxl.workbook import Workbook
import datetime

class ExcelUpgrader:
    def __init__(self, filename):
        self.filename = filename

    def upgrade(self):
        print(self.filename)

        # first open using xlrd
        book = xlrd.open_workbook(self.filename, formatting_info=True)
        index = 0
        nrows, ncols = 0, 0

        while nrows * ncols == 0:
            sheet = book.sheet_by_index(index)
            nrows = sheet.nrows
            ncols = sheet.ncols
            index += 1

        # prepare a xlsx sheet
        book1 = Workbook()
        sheet1 = book1.get_active_sheet()
        sheet1.title = 'hoja1'

        for row in range(1, nrows+1):
            for col in range(1, ncols+1):
                sheet1.cell(row=row, column=col).value = self._get_xlrd_cell_value(sheet.cell(row-1, col-1))

        book1.save(self.filename + "x")

    def _get_xlrd_cell_value(self,cell):
        value = cell.value
        if cell.ctype == xlrd.XL_CELL_DATE:
            value = datetime.datetime(*xlrd.xldate_as_tuple(value, 0))

        return value